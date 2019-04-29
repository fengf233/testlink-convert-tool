from openpyxl import load_workbook

XML_SUITE = r'''
<testsuite name="%(suite_name)s" >
<node_order><![CDATA[15]]></node_order>
<details><![CDATA[]]></details> 
%(xml_case)s
</testsuite>
'''
#%{sutie_name:'',xml_case:''}

XML_CASE = r'''
<testcase name="%(case_name)s">
	<node_order><![CDATA[1]]></node_order>
	<version><![CDATA[1]]></version>
    <summary><![CDATA[<p>%(summary)s</p>
]]></summary>
	<execution_type><![CDATA[1]]></execution_type>
	<importance><![CDATA[2]]></importance>
	<estimated_exec_duration></estimated_exec_duration>
	<status>1</status>
	<is_open>1</is_open>
	<active>1</active>
<steps>
%(xml_steps)s
</steps>
</testcase>
'''
#%{case_name:'',summary:'',xml_steps:''}

XML_STEPS = r'''
<step>
	<step_number><![CDATA[%(step_num)s]]></step_number>
	<actions><![CDATA[<p>%(step_text)s</p>
]]></actions>
	<expectedresults><![CDATA[<p>%(step_result)s</p>
]]></expectedresults>
	<execution_type><![CDATA[1]]></execution_type>
</step>
'''
#%(step_num)  %(step_text) %(step_result)

SUITE = r'''
<testsuite name="%s" >
<node_order><![CDATA[16]]></node_order>
<details><![CDATA[]]></details> 
'''
#可以理解为总的suite


def read_excel(col,start,end):
    a = {}
    for i in range(start,end):
        val = sheet[col+str(i)].value
        if val:
            a[i]=val
    return a

def read_excel2(col,start,end):
    a = {}
    for i in range(start,end):
        val = sheet[col+str(i)].value
        if val == None:
            #这里修改值为空时的赋值
            val = '-'
        a[i]=val
    return a

def suite_xml():
    suite = read_excel('a',1,max_row+1)
    suite_key = list(suite.keys())
    suite_key.append(max_row)
    s = []
    for i in range(0,len(suite_key)-1):
        cases = read_excel('b',suite_key[i],suite_key[i+1])
        s.append(XML_SUITE%{'suite_name':list(suite.values())[i],'xml_case':case_xml(cases,suite_key[i+1])})
    return s



def case_xml(case:dict,end)->str:
    case_key = list(case.keys())
    case_key.append(end)
    t = ''
    for i in range(0,len(case_key)-1):
        steps = read_excel2('d',case_key[i],case_key[i+1])
        results = read_excel2('e',case_key[i],case_key[i+1])
        summary = read_excel2('c',case_key[i],case_key[i]+1)
        new_steps ={}
        for x in steps.keys():
            new_steps[x]=[steps[x],results[x]]
        t= t+XML_CASE%{'case_name':list(case.values())[i],'summary':summary[case_key[i]],'xml_steps':steps_xml(new_steps)}
    return t

def steps_xml(steps:dict)->str:
    s = ''
    steps_key = list(steps.keys())
    for i in range(1,len(steps_key)+1):
        s = s + XML_STEPS%{'step_num':str(i),'step_text':steps[steps_key[i-1]][0],'step_result':steps[steps_key[i-1]][1]}
    return s



if __name__ == "__main__":
    #这里修改读取的xlsx
    wb = load_workbook('test.xlsx')
    sheet = wb.active
    max_row = sheet.max_row+1
    #这里修改保存xml的文件名称
    with open("test.xml",'w',encoding='utf-8') as f:
        f.write(r'''<?xml version="1.0" encoding="UTF-8"?>''')
        #这里改总的suite名称
        f.write(SUITE%'test_tmp')
        for a in suite_xml():
            f.write(a)
        f.write(r'''</testsuite>''')
    print('转化完成')